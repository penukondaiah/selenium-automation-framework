import inspect
import logging
import csv
from openpyxl import workbook, load_workbook



class Utils:
    def cust_logger(loglevel=logging.DEBUG):
        # set class/method name from where its called
        logger_name = inspect.stack()[1][3]
        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(loglevel)
        # create console handler or file handler and set the log level
        fh = logging.FileHandler("automation.log")
        # create formatter - how you want your logs to be formatted
        formatter = logging.Formatter("(%asctime)s - %(levelname)s - %(name)s : %(message)s")
        # add formatter to console or file handler
        fh.setFormatter(formatter)
        # add console handler to logger
        logger.addHandler(fh)
        return logger

    # def read_data_from_excel(file_name, sheet):
    #     datalist = []
    #     wb = load_workbook(filename=file_name)
    #     sh = wb[sheet]
    #
    #     row_ct = sh.max_row
    #     col_ct = sh.max_column
    #
    #     for i in range(2, row_ct + 1):
    #         row = []
    #         for j in range (1, col_ct + 1):
    #             row.append(sh.cell(row=i, column=j).value)
    #         datalist.append(row)
    #     return datalist


    # def read_data_from_csv(filename):
    #     #crete empty list
    #     datalist = []
    #
    #     #Open csv file
    #
    #     csvdata = open(filename, "r")
    #
    #     # create csv reaer
    #
    #     reader = csv.reader(csvdata)
    #
    #     #skip header
    #     next(reader)
    #     #Add csv rows to list
    #     for rows in reader:
    #         datalist.append(rows)
    #     return datalist

